#!/usr/bin/env python3
"""
Closeout <-> Debrief Reconciliation
===================================
Foxtrot Aviation Services

When a Commercial Closeout 2.0 webhook (JotForm 222916060752150) is received,
this module extracts the tails and services from the closeout and reconciles
them against the per-program Debrief workbooks
(SharePoint: DataHub > Power Flows > Debriefs > [Program] Debriefs.xlsx).

It checks BOTH directions for the closeout's date + location:
  * every debrief tail/service is reflected in the closeout, and
  * every closeout tail/service is reflected in the debriefs.

Any discrepancies are collected and, if present, an email is drafted via the
Claude API summarizing them.

------------------------------------------------------------------------------
CONFIG FLAGS  (the three decisions you can flip in one line each)
------------------------------------------------------------------------------
"""

import os
import re
import json
import datetime
from collections import defaultdict
import io
import time

from openpyxl import load_workbook

# ----- CONFIG -----------------------------------------------------------------

# (1) At CVG a closeout carries both PSA (field 27) and Envoy (field 45) aircraft.
#     True  -> route each fleet to its own debrief sheet and reconcile both.
#     False -> reconcile only the single primary program for the location.
ROUTE_MULTI_PROGRAM = True

# (2) Comparison strictness.
#     False -> tail + services: flag missing/extra tails AND service mismatches.
#     True  -> tail-level only: flag only missing/extra tails, ignore services.
TAIL_LEVEL_ONLY = False

# (3) Locations that must NOT be processed (skip the whole run if matched).
#     STL AD HOC is a distinct location string from STL; STL (GoJet) IS processed.
#     IAH was previously skipped; it is now reconciled via the Mesa debrief
#     workbook off field 298.
SKIP_LOCATIONS = {"DFW", "STL AD HOC"}

# ----- DEBRIEF SOURCE ---------------------------------------------------------
# DEBRIEF_SOURCE = "graph"  -> download workbooks from SharePoint via Microsoft
#                              Graph (production).
# DEBRIEF_SOURCE = "local"  -> read from local file paths (testing / mirrors).
DEBRIEF_SOURCE = os.environ.get("DEBRIEF_SOURCE", "graph")

# Local file paths (used when DEBRIEF_SOURCE == "local"). Overridable via env.
DEBRIEF_PATHS = {
    "GoJet": os.environ.get("GOJET_DEBRIEF",  "/mnt/user-data/uploads/GoJet_Debriefs.xlsx"),
    "PSA":   os.environ.get("PSA_DEBRIEF",    "/mnt/user-data/uploads/PSA_Debriefs.xlsx"),
    "Envoy": os.environ.get("ENVOY_DEBRIEF",  "/mnt/user-data/uploads/Envoy_Debriefs.xlsx"),
    "Mesa":  os.environ.get("MESA_DEBRIEF",   "/mnt/user-data/uploads/Mesa_Debriefs.xlsx"),
}

# SharePoint file paths (used when DEBRIEF_SOURCE == "graph"), relative to the
# root of the DataHub Shared Documents drive.
DEBRIEF_SP_PATHS = {
    "PSA":   "Power Flows/Debriefs/PSA Debriefs.xlsx",
    "Envoy": "Power Flows/Debriefs/Envoy Debriefs.xlsx",
    "GoJet": "Power Flows/Debriefs/GoJet Debriefs.xlsx",
    "Mesa":  "Power Flows/Debriefs/Mesa Debriefs.xlsx",
}

# Microsoft Graph / Entra credentials. Same Foxtrot Report Automation app used by
# the compliance trackers. Secret comes from the environment, never hard-coded.
GRAPH_TENANT_ID = os.environ.get("TENANT_ID", "ede0c57f-549f-4a90-9f8c-7ea130346f95")
GRAPH_CLIENT_ID = os.environ.get("CLIENT_ID", "58191600-ab56-4141-bff6-806805fcbff4")
GRAPH_CLIENT_SECRET = os.environ.get("CLIENT_SECRET", "")
GRAPH_DRIVE_ID = os.environ.get(
    "DRIVE_ID", "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU")

# Real-time race mitigation. Debriefs are completed during the shift and the
# closeout is submitted at end of shift, so overlap is not expected; retries are
# kept as a small safety net against transient Graph hiccups.
GRAPH_FETCH_RETRIES = int(os.environ.get("GRAPH_FETCH_RETRIES", "3"))
GRAPH_FETCH_DELAY_SEC = int(os.environ.get("GRAPH_FETCH_DELAY_SEC", "20"))

DEBRIEF_SHEETS = {"GoJet": "Sheet1", "PSA": "Debriefs", "Envoy": "Envoy General",
                  "Mesa": "Debriefs"}
# Envoy DFW debriefs are explicitly ignored (separate 'DFW' sheet AND DFW rows
# inside 'Envoy General' are both excluded).
ENVOY_IGNORE_LOCATIONS = {"DFW"}

# Email config — the drafted body is generic (no named addressee); these populate
# the to/from envelope. Overridable via EMAIL_TO env (comma-separated).
EMAIL_FROM = os.environ.get("EMAIL_FROM", "foxtrot.automation@foxtrotaviation.com")
EMAIL_TO   = [e.strip() for e in os.environ.get(
    "EMAIL_TO", "samuel.kosco@foxtrotaviation.com,clara.lana@foxtrotaviation.com,maren.pinpin@foxtrotaviation.com").split(",") if e.strip()]
CLAUDE_MODEL = "claude-opus-4-7"
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# IAH dispatch email — a separate per-closeout email that mirrors the existing
# nightly IAH dispatch David Blatt sends. Sent in addition to the reconciliation
# email whenever location == IAH. All three values are overridable via env so the
# sender/recipients can be rotated without code changes.
IAH_DISPATCH_SENDER_NAME  = os.environ.get("IAH_DISPATCH_SENDER_NAME",  "Renel Anthony")
IAH_DISPATCH_SENDER_EMAIL = os.environ.get("IAH_DISPATCH_SENDER_EMAIL", "renel.anthony@foxtrotaviation.com")
IAH_DISPATCH_RECIPIENTS   = [e.strip() for e in os.environ.get(
    "IAH_DISPATCH_RECIPIENTS",
    "samuel.kosco@foxtrotaviation.com,maren.pinpin@foxtrotaviation.com,david.blatt@foxtrotaviation.com,renel.anthony@foxtrotaviation.com,robert.murillo@foxtrotaviation.com,heidi.cromer@foxtrotaviation.com,daniel.digiambattista@foxtrotaviation.com,anthony.pentz@foxtrotaviation.com,clara.lana@foxtrotaviation.com,chris.stump@foxtrotaviation.com,nicholas.thomas@foxtrotaviation.com,jessica.clapper@foxtrotaviation.com,russell.dozier@mesa-air.com,shayla.ortiz@mesa-air.com,eric.nation@mesa-air.com,andy.jamison@united.com").split(",") if e.strip()]

# IAH dispatch is once-per-day: the first IAH closeout for a given date (field 4)
# sends the dispatch email; later closeouts for that same date are resubmissions
# (edited values) and must NOT re-send the dispatch — but the reconciliation
# ("comparison") email still runs on every submission. We remember which dates
# already triggered a dispatch in a small JSON file on the DataHub drive, so the
# memory survives across GitHub Actions runs. A bounded recent-date history (not
# just the single latest date) keeps this correct even if submissions arrive out
# of order or span multiple days.
IAH_DISPATCH_STATE_PATH = os.environ.get(
    "IAH_DISPATCH_STATE_PATH",
    "Power Flows/Commercial Closeout/iah_dispatch_state.json")
IAH_DISPATCH_STATE_KEEP = int(os.environ.get("IAH_DISPATCH_STATE_KEEP", "60"))

# Discrepancy records → Power Automate. For every discrepancy found during
# reconciliation, a JSON packet is POSTed as its own HTTP request to a Power
# Automate "When an HTTP request is received" flow, which appends a row to a
# compliance Excel table so analysts can mark the outcome. This mirrors the
# "PSA Tail Add" flow on the compliance tracker (one POST per item → one Add a
# row). The trigger URL carries its own SAS-style signature; store it as a
# GitHub secret / workflow env var rather than committing it. When unset,
# posting is skipped but the records are still logged to the Actions console.
DISCREPANCY_WEBHOOK_URL = os.environ.get("DISCREPANCY_WEBHOOK_URL", "")

# ----- SERVICE VOCABULARY NORMALIZATION --------------------------------------
# Both systems get mapped to a single canonical code set so they can be compared.
# Canonical codes: I, Ex, CC, DSC, CE, ED1, ED2, ED3, ED4, IHC, RON, LAV,
#                  plus Mesa-specific: EC, ED, FCD, ESS, Flight Deck.
# (Mesa uses EC/ED distinct from Ex/ED1-4 because the JotForm field 298 multi-
# select exposes the Mesa abbreviations directly. Reconciliation is per-fleet so
# the separate codes never compare across fleets.)

# Closeout 'Service Performed' tokens -> canonical
CLOSEOUT_TOKEN_MAP = {
    "I": "I", "INT": "I", "INTERIOR": "I",
    "EX": "Ex", "E": "Ex", "EXTERIOR": "Ex",
    "CC": "CC",
    "DSC": "DSC",
    "CE": "CE",
    "ED1": "ED1", "ED2": "ED2", "ED3": "ED3", "ED4": "ED4",
    "IHC": "IHC",
    "RON": "RON",
    "LAV": "LAV",
    # Mesa (IAH) field 298 abbreviations
    "EC": "EC",
    "ED": "ED",
    "FCD": "FCD",
    "ESS": "ESS",
    "FLIGHT DECK": "Flight Deck",
}

# Debrief column header -> canonical (only service columns; cols 0-3 are
# Date/Name/Location/Tail and the last is Sub ID — except the Mesa workbook,
# which has no Location column so Tail sits at col 2).
DEBRIEF_COL_MAP = {
    "Interior Clean (I)": "I",
    "Exterior Clean (E)": "Ex",
    "Cockpit Cleaning (CC)": "CC",
    "Deep Seat Clean (DSC)": "DSC",
    "Carpet Extraction (CE)": "CE",
    "Exterior Detail (ED1)": "ED1",
    "Exterior Detail (ED2)": "ED2",
    "Exterior Detail (ED3)": "ED3",
    "Exterior Detail (ED4)": "ED4",
    "Exterior Detail #1 (ED1)": "ED1",
    "Exterior Detail #2 (ED2)": "ED2",
    "Lav Tank Pressure Washing": "LAV",
    "Interior Heavy Clean (IHC)": "IHC",
    "IHC": "IHC",
    "RON Clean (RON)": "RON",
    # Mesa-specific columns
    "Exterior Clean (EC)": "EC",
    "Exterior Detail (ED)": "ED",
    "Fleet Campaign Decal (FCD)": "FCD",
    "Disinfection (ESS)": "ESS",
    "Detailed Flight Deck Clean (Flight Deck)": "Flight Deck",
}


def _canon_closeout_service(raw_service_string):
    """A closeout 'Service Performed' value -> set of canonical codes.

    Handles three shapes:
      * newline-delimited PSA/Envoy tokens: "I\\nEx\\nCC\\nDSC\\nCE"
      * bare "RON"
      * GoJet language with code in parens:
            "Gojet CLN-02 (IHC)\\nGojet CLN-06 (ED1)\\nGojet CLN-04 (CE)"
    """
    codes = set()
    if not raw_service_string:
        return codes
    for line in str(raw_service_string).replace("\r", "").split("\n"):
        line = line.strip()
        if not line:
            continue
        # GoJet: prefer the code inside parentheses if present
        paren = re.search(r"\(([^)]+)\)", line)
        token = paren.group(1) if paren else line
        key = token.strip().upper()
        if key in CLOSEOUT_TOKEN_MAP:
            codes.add(CLOSEOUT_TOKEN_MAP[key])
    return codes


# ----- CLOSEOUT EXTRACTION ----------------------------------------------------
# Verified field map: 6=Location, 4=Date, 3=Submitter,
# 281=GoJet (tail key 'Tail Number'), 27=PSA ('Dropdown'),
# 45=Envoy ('Tail Number'), 298=Mesa/IAH ('Tail Number', services under 'Services').
# Mesa's per-row object also carries Gate/Start/Complete/Notes from the JotForm
# configurable list; those are ignored by the reconciler.

FLEET_FIELDS = [
    ("GoJet", "281", "Tail Number", "Service Performed"),
    ("PSA",   "27",  "Dropdown",    "Service Performed"),
    ("Envoy", "45",  "Tail Number", "Service Performed"),
    ("Mesa",  "298", "Tail Number", "Services"),
]


def _parse_array(field):
    if not field or not str(field).strip():
        return []
    try:
        return json.loads(field)
    except (json.JSONDecodeError, TypeError):
        return []


def extract_closeout(body):
    """Return dict: location, date(datetime.date), submitter, and
    fleets -> {tail: set(canonical services)}."""
    loc = (body.get("6") or "").strip()
    date_raw = (body.get("4") or "").strip()
    try:
        date = datetime.datetime.strptime(date_raw, "%Y-%m-%d").date()
    except ValueError:
        date = None

    fleets = defaultdict(lambda: defaultdict(set))
    for fleet, fid, tail_key, svc_key in FLEET_FIELDS:
        for row in _parse_array(body.get(fid)):
            tail = (row.get(tail_key) or "").strip().upper()
            if not tail:
                continue
            fleets[fleet][tail] |= _canon_closeout_service(row.get(svc_key))
    return {"location": loc, "date": date, "submitter": (body.get("3") or "").strip(),
            "fleets": {k: dict(v) for k, v in fleets.items()}}


# ----- DEBRIEF EXTRACTION -----------------------------------------------------

def _location_matches(debrief_loc, closeout_loc, fleet):
    """Match a debrief location against a closeout location on the airport code.

    Either side may carry a '-Program' suffix (e.g. the closeout sends 'SGF-Envoy'
    while the debrief stores bare 'SGF', or vice-versa with 'DCA-PSA'). Strip the
    suffix from BOTH sides and compare the bare airport code."""
    if debrief_loc is None:
        return False
    db_base = str(debrief_loc).strip().upper().split("-")[0]   # 'DCA-PSA' -> 'DCA'
    co_base = str(closeout_loc).strip().upper().split("-")[0]  # 'SGF-Envoy' -> 'SGF'
    return db_base == co_base


# ----- MICROSOFT GRAPH (SharePoint download) ---------------------------------

class _NonRetryableGraphError(RuntimeError):
    """Graph error that should fail immediately (e.g. 404 path not found)."""


_graph_token_cache = {"token": None, "expires_at": 0.0}


def _graph_token():
    """Obtain (and cache) an app-only Graph access token via client credentials.
    Uses the Foxtrot Report Automation Entra app — same one the trackers use."""
    import requests  # imported lazily so local-mode runs need no network deps

    now = time.time()
    if _graph_token_cache["token"] and now < _graph_token_cache["expires_at"]:
        return _graph_token_cache["token"]

    if not GRAPH_CLIENT_SECRET:
        raise RuntimeError("CLIENT_SECRET not set — cannot authenticate to Graph")

    url = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "client_id": GRAPH_CLIENT_ID,
        "client_secret": GRAPH_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    _graph_token_cache["token"] = data["access_token"]
    # refresh a minute early to avoid edge expiry
    _graph_token_cache["expires_at"] = now + int(data.get("expires_in", 3600)) - 60
    return _graph_token_cache["token"]


def _graph_download_workbook_bytes(sp_path):
    """Download a workbook's bytes from the DataHub drive by its path.
    Retries to mitigate the real-time race with Power Automate's debrief writes."""
    import requests

    if not GRAPH_DRIVE_ID:
        raise RuntimeError("DRIVE_ID not set — cannot locate the SharePoint drive")

    # URL-encode the path segments but keep the slashes
    encoded = "/".join(requests.utils.quote(seg) for seg in sp_path.split("/"))
    url = (f"https://graph.microsoft.com/v1.0/drives/{GRAPH_DRIVE_ID}"
           f"/root:/{encoded}:/content")

    last_err = None
    for attempt in range(1, GRAPH_FETCH_RETRIES + 1):
        try:
            headers = {"Authorization": f"Bearer {_graph_token()}"}
            resp = requests.get(url, headers=headers, timeout=60)
            if resp.status_code == 200:
                return resp.content
            if resp.status_code == 404:
                # Genuine path error — don't retry; re-raise past the retry handler
                raise _NonRetryableGraphError(
                    f"404 Not Found downloading '{sp_path}'. Check the SharePoint "
                    f"path (and confirm the GoJet path if this is STL).")
            resp.raise_for_status()
        except _NonRetryableGraphError:
            raise  # fail fast, no retry
        except Exception as e:  # noqa: BLE001 — transient; surface after final attempt
            last_err = e
            if attempt < GRAPH_FETCH_RETRIES:
                time.sleep(GRAPH_FETCH_DELAY_SEC)
    raise RuntimeError(f"Failed to download '{sp_path}' after "
                       f"{GRAPH_FETCH_RETRIES} attempts: {last_err}")


def _graph_get_json(sp_path):
    """GET a small JSON file from the DataHub drive by path. Returns the parsed
    object, or None if the file doesn't exist yet (404) or can't be read. State
    reads are best-effort — a failure here must never block a dispatch."""
    import requests

    if not GRAPH_DRIVE_ID:
        return None
    encoded = "/".join(requests.utils.quote(seg) for seg in sp_path.split("/"))
    url = (f"https://graph.microsoft.com/v1.0/drives/{GRAPH_DRIVE_ID}"
           f"/root:/{encoded}:/content")
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {_graph_token()}"},
                            timeout=30)
        if resp.status_code == 404:
            return None  # no state yet — first run
        resp.raise_for_status()
        return resp.json()
    except Exception as e:  # noqa: BLE001 — best-effort; surface but don't raise
        print(f"[state read failed ({sp_path}): {e}]", flush=True)
        return None


def _graph_put_json(sp_path, obj):
    """PUT a small JSON file to the DataHub drive by path, creating or
    overwriting it. Raises on failure so the caller can log it."""
    import requests

    if not GRAPH_DRIVE_ID:
        raise RuntimeError("DRIVE_ID not set — cannot write state file")
    encoded = "/".join(requests.utils.quote(seg) for seg in sp_path.split("/"))
    url = (f"https://graph.microsoft.com/v1.0/drives/{GRAPH_DRIVE_ID}"
           f"/root:/{encoded}:/content")
    resp = requests.put(url, headers={
        "Authorization": f"Bearer {_graph_token()}",
        "Content-Type": "application/json",
    }, data=json.dumps(obj).encode("utf-8"), timeout=30)
    resp.raise_for_status()


def _open_debrief_workbook(fleet):
    """Return an openpyxl workbook for the fleet, from Graph or local disk."""
    if DEBRIEF_SOURCE == "graph":
        sp_path = DEBRIEF_SP_PATHS[fleet]
        content = _graph_download_workbook_bytes(sp_path)
        return load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    # local
    return load_workbook(DEBRIEF_PATHS[fleet], read_only=True, data_only=True)


def load_debrief_day(fleet, closeout_loc, date):
    """Return {tail: set(canonical services)} for the given fleet/location/date.

    PSA/Envoy/GoJet workbooks are laid out Date/Name/Location/Tail at cols 0-3.
    The Mesa workbook has no Location column — its layout is Date/Name/Tail at
    cols 0-2 — and every row is implicitly IAH, so no location filtering runs."""
    sheet = DEBRIEF_SHEETS[fleet]
    wb = _open_debrief_workbook(fleet)
    ws = wb[sheet]
    header = next(ws.iter_rows(max_row=1, values_only=True))
    # Map service columns by canonical code
    svc_cols = {i: DEBRIEF_COL_MAP[h] for i, h in enumerate(header)
                if h in DEBRIEF_COL_MAP}

    has_location = (fleet != "Mesa")
    tail_idx = 3 if has_location else 2

    result = defaultdict(set)
    row_counts = defaultdict(int)            # tail -> number of debrief rows today
    occurrences = defaultdict(list)          # tail -> [sorted services] per row
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = r[0]
        if isinstance(d, datetime.datetime):
            d = d.date()
        if d != date:
            continue
        if has_location:
            loc = r[2]
            if fleet == "Envoy" and str(loc).strip().upper() in ENVOY_IGNORE_LOCATIONS:
                continue  # ignore DFW Envoy rows
            if not _location_matches(loc, closeout_loc, fleet):
                continue
        tail = r[tail_idx]
        tail = (str(tail).strip().upper() if tail else "")
        if not tail:
            continue
        row_services = set()
        for i, code in svc_cols.items():
            val = r[i]
            if isinstance(val, str) and val.strip().lower().startswith("yes"):
                row_services.add(code)
        result[tail] |= row_services
        row_counts[tail] += 1
        occurrences[tail].append(sorted(row_services))
    wb.close()

    # A tail appearing on more than one debrief row for this date/location is a
    # double submission. Services are still collapsed to a single set for
    # reconciliation, but each duplicate is surfaced separately so it can be
    # flagged in the email. occurrences carries the services from each submission
    # so the note can show whether the duplicates are identical.
    duplicates = [
        {"tail": tail, "count": row_counts[tail], "occurrences": occurrences[tail]}
        for tail in sorted(row_counts) if row_counts[tail] > 1
    ]
    return dict(result), duplicates


# ----- TYPO DETECTION ---------------------------------------------------------
# Closeout tails are free-typed (text box); debrief tails come from a dropdown
# and are treated as the source of truth. When a closeout tail has no debrief
# match and vice-versa, check whether the pair is "close enough" to be a typo:
# exactly one character dropped/changed, or two adjacent characters swapped.
# (Damerau-Levenshtein distance == 1, with substring length difference <= 1.)

def _is_one_edit_away(a, b):
    """True if b can be reached from a by exactly one of:
    single deletion, single substitution, or one adjacent transposition.
    (Insertion into a == deletion from b, so this is symmetric.)"""
    if a == b:
        return False  # identical is a match, not a typo
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return False

    if la == lb:
        # substitution (1 differing position) OR adjacent transposition
        diffs = [i for i in range(la) if a[i] != b[i]]
        if len(diffs) == 1:
            return True  # single substitution
        if len(diffs) == 2:
            i, j = diffs
            return j == i + 1 and a[i] == b[j] and a[j] == b[i]  # adjacent swap
        return False

    # length differs by exactly 1 -> single deletion (drop one char from longer)
    longer, shorter = (a, b) if la > lb else (b, a)
    i = j = 0
    skipped = False
    while i < len(longer) and j < len(shorter):
        if longer[i] == shorter[j]:
            i += 1
            j += 1
        elif skipped:
            return False
        else:
            skipped = True
            i += 1  # skip one char in the longer string
    return True  # at most one skip used


def _find_typo_pairs(co_only, db_only):
    """Match unmatched closeout tails against unmatched debrief tails.
    Returns (pairs, leftover_co, leftover_db) where pairs is a list of
    (closeout_tail, debrief_tail). Only pairs that are UNAMBIGUOUS in both
    directions (each is the other's sole near-match) are reported as typos;
    ambiguous or unmatched tails fall through as genuine missing tails."""
    co_list, db_list = list(co_only), list(db_only)
    # candidate near-matches each way
    co_cands = {c: [d for d in db_list if _is_one_edit_away(c, d)] for c in co_list}
    db_cands = {d: [c for c in co_list if _is_one_edit_away(c, d)] for d in db_list}

    pairs, used_co, used_db = [], set(), set()
    for c in co_list:
        cands = co_cands[c]
        if len(cands) == 1:
            d = cands[0]
            # require it to be mutual and unambiguous from the debrief side too
            if len(db_cands[d]) == 1 and db_cands[d][0] == c:
                pairs.append((c, d))
                used_co.add(c)
                used_db.add(d)
    leftover_co = [c for c in co_list if c not in used_co]
    leftover_db = [d for d in db_list if d not in used_db]
    return pairs, leftover_co, leftover_db


# ----- RECONCILIATION ---------------------------------------------------------

def reconcile_fleet(fleet, closeout_tails, debrief_tails):
    """Compare one fleet's closeout tails vs debrief tails for the day.
    Returns a dict of discrepancy lists."""
    disc = {
        "missing_in_debrief": [],   # on closeout, not in debrief, no typo match
        "missing_in_closeout": [],  # in debrief, not on closeout, no typo match
        "probable_typos": [],       # closeout tail ~ debrief tail (1-edit apart)
        "service_mismatches": [],   # tail matched (exact or typo), services differ
    }
    co_set, db_set = set(closeout_tails), set(debrief_tails)
    co_only, db_only = co_set - db_set, db_set - co_set

    # Identify probable typos among the unmatched tails on each side.
    typo_pairs, co_left, db_left = _find_typo_pairs(co_only, db_only)

    # Genuine missing tails (no plausible typo partner)
    for tail in sorted(co_left):
        disc["missing_in_debrief"].append(
            {"tail": tail, "closeout_services": sorted(closeout_tails[tail])})
    for tail in sorted(db_left):
        disc["missing_in_closeout"].append(
            {"tail": tail, "debrief_services": sorted(debrief_tails[tail])})

    # Probable-typo pairs: still reported as an error, flagged as likely a
    # closeout typo. Because they are most likely the same aircraft, compare
    # their services too (unless tail-level-only mode).
    for co_tail, db_tail in sorted(typo_pairs):
        entry = {
            "closeout_tail": co_tail,
            "debrief_tail": db_tail,
            "closeout_services": sorted(closeout_tails[co_tail]),
            "debrief_services": sorted(debrief_tails[db_tail]),
        }
        if not TAIL_LEVEL_ONLY:
            co_svc, db_svc = closeout_tails[co_tail], debrief_tails[db_tail]
            entry["service_match"] = (co_svc == db_svc)
            entry["on_closeout_only"] = sorted(co_svc - db_svc)
            entry["on_debrief_only"] = sorted(db_svc - co_svc)
        disc["probable_typos"].append(entry)

    # Exact-match tails: compare services
    if not TAIL_LEVEL_ONLY:
        for tail in sorted(co_set & db_set):
            co_svc, db_svc = closeout_tails[tail], debrief_tails[tail]
            if co_svc != db_svc:
                disc["service_mismatches"].append({
                    "tail": tail,
                    "on_closeout_only": sorted(co_svc - db_svc),
                    "on_debrief_only": sorted(db_svc - co_svc),
                    "closeout_services": sorted(co_svc),
                    "debrief_services": sorted(db_svc),
                })
    return disc


def reconcile(body):
    """Top-level entry point. Returns a structured report dict.
    report['skipped'] is set when the location is in SKIP_LOCATIONS."""
    co = extract_closeout(body)
    loc, date = co["location"], co["date"]

    # Skip check compares the airport code with any '-Program' suffix removed, so
    # 'DFW-Envoy'/'IAH-PSA' are caught. Multi-word skips like 'STL AD HOC' have no
    # dash, so their base is themselves and a plain 'STL' will NOT match them.
    loc_base = loc.strip().upper().split("-")[0]
    skip_bases = {s.upper().split("-")[0] for s in SKIP_LOCATIONS}
    if loc_base in skip_bases:
        return {"skipped": True, "reason": f"Location '{loc}' is in SKIP_LOCATIONS",
                "location": loc, "date": str(date)}
    if date is None:
        return {"skipped": True, "reason": "Closeout date could not be parsed",
                "location": loc, "date": co["date"]}

    fleets_present = {f: t for f, t in co["fleets"].items() if t}
    if not ROUTE_MULTI_PROGRAM and len(fleets_present) > 1:
        # keep only the largest fleet
        primary = max(fleets_present, key=lambda f: len(fleets_present[f]))
        fleets_present = {primary: fleets_present[primary]}

    per_fleet = {}
    has_any = False
    has_dupes = False
    for fleet, co_tails in fleets_present.items():
        db_tails, db_duplicates = load_debrief_day(fleet, loc, date)
        d = reconcile_fleet(fleet, co_tails, db_tails)
        per_fleet[fleet] = {
            "closeout_tail_count": len(co_tails),
            "debrief_tail_count": len(db_tails),
            "discrepancies": d,
            # debrief tails submitted on more than one debrief today (double debriefs)
            "duplicate_debriefs": db_duplicates,
            # tail -> sorted service codes, as serviced per the closeout
            "serviced": {t: sorted(s) for t, s in co_tails.items()},
        }
        if any(d.values()):
            has_any = True
        if db_duplicates:
            has_dupes = True

    # has_duplicates is tracked separately from has_discrepancies so a run whose
    # ONLY finding is a double debrief still routes to the clean-email path (with
    # a flag) rather than the Claude discrepancy narrative.
    return {"skipped": False, "location": loc, "date": str(date),
            "submitter": co["submitter"], "fleets": per_fleet,
            "has_discrepancies": has_any, "has_duplicates": has_dupes}


# ----- EMAIL DRAFTING (Claude API) -------------------------------------------

def draft_discrepancy_email(report):
    """Call the Claude API to draft a short internal email summarizing the
    discrepancies. Returns {'subject':..., 'body':...} or None on failure."""
    if report.get("skipped") or not report.get("has_discrepancies"):
        return None

    try:
        import anthropic
    except ImportError:
        raise RuntimeError("pip install anthropic to enable email drafting")

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    prompt = f"""You are drafting a brief INTERNAL email flagging reconciliation discrepancies between a nightly Commercial Closeout and the program debrief tracker for Foxtrot Aviation Services.

Here is the structured discrepancy report (JSON):

{json.dumps(report, indent=2)}

Write a concise, professional email. Requirements:
- Subject line on the first line prefixed with "Subject: ".
- Do NOT include a salutation/greeting or a named addressee (no "Hi", no "Dear ___") — the recipient list is set separately. Begin directly with the summary sentence.
- Open with one sentence stating the location, date, and that discrepancies were found.
- For each fleet, use a short labeled section. List:
    * Tails on the closeout but missing from the debrief.
    * Tails in the debrief but missing from the closeout.
    * Tails where the service lists differ (name the differing service codes on each side).
    * Probable typos: a closeout tail that does not match any debrief tail but is one character off from a debrief tail that is otherwise unmatched. Report these as an error, BUT explicitly note it is most likely a typo on the closeout, showing both the closeout-entered tail and the likely-correct debrief tail. If the matched services also differ, mention that too.
    * Double debriefs: under each fleet's "duplicate_debriefs", any tail that was submitted on more than one debrief for this date. List each as a flagged discrepancy: name the tail, state how many debriefs it appeared on, and show the services recorded on each submission. Call out whether the duplicate submissions are identical or differ.
- Keep it scannable (short lines / simple bullets). No filler and no closing pleasantries; end the body after the last item with no sign-off.
- Do not invent any tails or services beyond what is in the JSON."""

    msg = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=1200,
        messages=[{"role": "user", "content": prompt}],
    )
    text = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text").strip()

    subject = f"Closeout/Debrief discrepancies — {report['location']} {report['date']}"
    body = text
    if text.lower().startswith("subject:"):
        first, _, rest = text.partition("\n")
        subject = first.split(":", 1)[1].strip()
        body = rest.strip()
    return {"subject": subject, "body": body, "to": EMAIL_TO, "from": EMAIL_FROM}


def send_email_via_graph(email):
    """Send a drafted email through Microsoft Graph as EMAIL_FROM.
    Reuses the same Entra app/token as the SharePoint download (sendMail scope).
    `email` is the dict returned by draft_discrepancy_email."""
    import requests

    if not email or not email.get("to"):
        raise RuntimeError("No recipients set (EMAIL_TO) — cannot send.")

    url = (f"https://graph.microsoft.com/v1.0/users/"
           f"{requests.utils.quote(email['from'])}/sendMail")
    payload = {
        "message": {
            "subject": email["subject"],
            "body": {"contentType": email.get("content_type", "Text"),
                     "content": email["body"]},
            "toRecipients": [{"emailAddress": {"address": a}} for a in email["to"]],
        },
        "saveToSentItems": True,
    }
    resp = requests.post(url, headers={
        "Authorization": f"Bearer {_graph_token()}",
        "Content-Type": "application/json",
    }, json=payload, timeout=30)
    if resp.status_code not in (200, 202):
        raise RuntimeError(f"sendMail failed ({resp.status_code}): {resp.text[:300]}")
    return True


def build_clean_email(report):
    """Build the no-discrepancy confirmation email deterministically (no API).
    Includes an HTML table of every tail serviced and its services."""
    import html

    loc = report["location"]
    subject = f"{loc} closeout — no discrepancies ({report['date']})"

    intro = (f"{loc} had no discrepancies between the services on the closeout "
             f"and the debriefs.")

    # Collect rows across all fleets: (tail, fleet, services-string)
    rows = []
    for fleet, info in report["fleets"].items():
        for tail, services in sorted(info.get("serviced", {}).items()):
            rows.append((tail, fleet, ", ".join(services) if services else "—"))

    th = ('style="text-align:left;padding:6px 12px;border:1px solid #ccc;'
          'background:#1F3864;color:#fff;"')
    td = 'style="padding:6px 12px;border:1px solid #ccc;"'
    header = (f"<tr><th {th}>Tail Number</th><th {th}>Fleet</th>"
              f"<th {th}>Services</th></tr>")
    body_rows = "".join(
        f"<tr><td {td}>{html.escape(t)}</td><td {td}>{html.escape(f)}</td>"
        f"<td {td}>{html.escape(s)}</td></tr>" for t, f, s in rows)
    table = (f'<table style="border-collapse:collapse;font-family:Arial,'
             f'sans-serif;font-size:13px;">{header}{body_rows}</table>')

    # Double-debrief flag. This is the no-discrepancy path, so when a duplicate is
    # present it is the only finding: mark the subject and append a bold note
    # below the table describing each double submission.
    dup_entries = [(fleet, dup)
                   for fleet, info in report["fleets"].items()
                   for dup in info.get("duplicate_debriefs", [])]
    dup_note = ""
    if dup_entries:
        subject += " (Double Debrief Flagged)"
        items = []
        for fleet, dup in dup_entries:
            occ = "; ".join("[" + ", ".join(o) + "]" if o else "[none]"
                            for o in dup["occurrences"])
            items.append(
                f"<li>{html.escape(dup['tail'])} ({html.escape(fleet)}) — submitted on "
                f"{dup['count']} debriefs for this date "
                f"(services per submission: {html.escape(occ)})</li>")
        dup_note = (
            '<p style="font-weight:bold;color:#b00020;margin-top:16px;">'
            'Double debrief flagged — the following tail(s) were submitted on more '
            'than one debrief for this date:</p>'
            f'<ul style="font-weight:bold;color:#b00020;">{"".join(items)}</ul>')

    body_html = (f'<div style="font-family:Arial,sans-serif;font-size:14px;">'
                 f'<p>{html.escape(intro)}</p>'
                 f'<p><b>Aircraft serviced ({len(rows)}):</b></p>'
                 f'{table}{dup_note}</div>')

    return {"subject": subject, "body": body_html, "content_type": "HTML",
            "to": EMAIL_TO, "from": EMAIL_FROM}


# ----- IAH DISPATCH EMAIL ----------------------------------------------------
# Mirrors the nightly dispatch David Blatt sends. Built directly from the raw
# closeout body (fields 4, 298, 299) rather than from the reconcile report,
# because the report drops Gate/Start/Complete/Notes from field 298 and never
# sees field 299 at all. Sent in addition to the reconciliation email whenever
# the closeout location is IAH.

def _format_iah_time(raw):
    """Convert a JotForm time string like '05:49 PM' to 24-hour 'H:MM'.
    Returns the raw value verbatim if it can't be parsed."""
    s = str(raw or "").strip()
    if not s:
        return ""
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M"):
        try:
            t = datetime.datetime.strptime(s, fmt)
            return f"{t.hour}:{t.minute:02d}"
        except ValueError:
            continue
    return s


def _build_iah_services_table(rows_298):
    """HTML table from field 298 rows. Services rendered as slash-delimited."""
    import html as _html
    th = ('style="background:#1F3864;color:#ffffff;font-weight:bold;'
          'text-decoration:underline;padding:8px 12px;border:1px solid #000;'
          'text-align:center;"')
    td = ('style="padding:6px 12px;border:1px solid #000;text-align:center;"')
    header = (f"<tr><th {th}>TAIL</th><th {th}>GATE</th><th {th}>SERVICES</th>"
              f"<th {th}>START</th><th {th}>COMPLETE</th><th {th}>NOTES</th></tr>")

    body_rows = []
    for row in rows_298:
        tail = str(row.get("Tail Number") or "").strip()
        gate = str(row.get("Gate") or "").strip()
        services_raw = str(row.get("Services") or "")
        services = [s.strip() for s in services_raw.replace("\r", "").split("\n") if s.strip()]
        services_str = "/".join(services)
        start = _format_iah_time(row.get("Start"))
        complete = _format_iah_time(row.get("Complete"))
        notes = str(row.get("Notes") or "").strip()
        body_rows.append(
            f"<tr><td {td}>{_html.escape(tail)}</td>"
            f"<td {td}>{_html.escape(gate)}</td>"
            f"<td {td}>{_html.escape(services_str)}</td>"
            f"<td {td}>{_html.escape(start)}</td>"
            f"<td {td}>{_html.escape(complete)}</td>"
            f"<td {td}>{_html.escape(notes)}</td></tr>"
        )

    return ('<table style="border-collapse:collapse;font-family:Arial,sans-serif;'
            f'font-size:13px;margin:12px 0;">{header}{"".join(body_rows)}</table>')


def _build_iah_observations_table(rows_299):
    """HTML table from field 299 rows. Returns None if the list is empty."""
    import html as _html
    if not rows_299:
        return None
    th = ('style="background:#9BD18A;color:#000000;font-weight:bold;'
          'text-decoration:underline;padding:8px 12px;border:1px solid #000;'
          'text-align:center;"')
    td = ('style="padding:6px 12px;border:1px solid #000;text-align:center;"')
    tail_td = ('style="padding:6px 12px;border:1px solid #000;text-align:center;'
               'width:120px;"')
    header = (f'<tr><th {th} width="120">TAIL</th>'
              f'<th {th}>OBSERVATIONS</th></tr>')

    body_rows = []
    for row in rows_299:
        tail = str(row.get("Tail") or "").strip()
        observation = str(row.get("Observation") or "").strip()
        body_rows.append(
            f"<tr><td {tail_td}>{_html.escape(tail)}</td>"
            f"<td {td}>{_html.escape(observation)}</td></tr>"
        )

    return ('<table style="border-collapse:collapse;font-family:Arial,sans-serif;'
            f'font-size:13px;margin:12px 0;">{header}{"".join(body_rows)}</table>')


def build_iah_dispatch_email(body, sender_name=None, sender_email=None):
    """Build the IAH per-closeout dispatch email from the raw JotForm body.
    Reads fields 4 (date), 298 (Mesa fleet rows), 299 (observations).

    sender_name / sender_email override the module defaults so the dispatch can
    be sent as the submitter (constructed from field 3) and fall back to the
    configured defaults if that send fails. Both fall through to the
    IAH_DISPATCH_* constants when omitted."""
    import html as _html

    name = sender_name or IAH_DISPATCH_SENDER_NAME
    email_from = sender_email or IAH_DISPATCH_SENDER_EMAIL

    date_raw = (body.get("4") or "").strip()
    try:
        date_obj = datetime.datetime.strptime(date_raw, "%Y-%m-%d").date()
        date_str = date_obj.strftime("%B %d, %Y")
        # Strip leading zero off day-of-month for a more natural rendering
        date_str = date_str.replace(f" 0{date_obj.day}, ", f" {date_obj.day}, ")
    except ValueError:
        date_str = date_raw

    rows_298 = _parse_array(body.get("298"))
    rows_299 = _parse_array(body.get("299"))
    aircraft_count = len(rows_298)

    services_table = _build_iah_services_table(rows_298)
    observations_table = _build_iah_observations_table(rows_299)
    observations_block = observations_table if observations_table else (
        '<p style="font-family:Arial,sans-serif;font-size:14px;">'
        'No observations on Mesa-Air aircraft today.</p>'
    )

    subject = (f"IAH Dispatch Foxtrot Aviation Services: "
               f"Mesa-Air Closeout {date_str}")

    body_html = (
        f'<div style="font-family:Arial,sans-serif;font-size:14px;">'
        f'<p>Hello everyone,</p>'
        f'<p>Please review the dispatch information for details regarding the '
        f'completed services and observations from {_html.escape(date_str)}. '
        f'The team successfully serviced {aircraft_count} Mesa-Air aircraft.</p>'
        f'{services_table}'
        f'{observations_block}'
        f'<p>Thanks,</p>'
        f'<p>{_html.escape(name)}<br>'
        f'Foxtrot Aviation Services</p>'
        f'</div>'
    )

    return {"subject": subject, "body": body_html, "content_type": "HTML",
            "to": IAH_DISPATCH_RECIPIENTS, "from": email_from}


# ----- DISCREPANCY RECORDS (Power Automate webhook) --------------------------
# For every discrepancy in the report, emit one flat record and POST it to a
# Power Automate flow that appends a row to a compliance Excel table. Each
# record carries Location, Tail Number, Date, Discrepancy, and Program. The
# Discrepancy string is human-readable and self-explaining so an analyst can act
# on it without opening the closeout:
#   * "Missing from Debrief"  — on the closeout, no matching debrief tail
#   * "Missing from Closeout" — in the debrief, no matching closeout tail
#   * "Service Mismatch (...)"— matched tail, service lists differ
#   * "Possible Typo (...)"   — closeout tail is one edit off an unmatched debrief tail
#   * "Double Debrief"        — tail submitted on more than one debrief that day
# ("Misdated" is a planned future value; it is not produced by the current
# reconciliation logic.)

# Field names match the JSON the Power Automate HTTP trigger expects (and, in
# turn, the Excel table columns the flow writes).
_DISC_LOCATION = "Location"
_DISC_TAIL     = "Tail Number"
_DISC_DATE     = "Date"
_DISC_VALUE    = "Discrepancy"
_DISC_PROGRAM  = "Program"


def _fmt_service_delta(on_closeout_only, on_debrief_only):
    """Render the two-sided service difference for a mismatch/typo, e.g.
    'closeout-only: CC, DSC; debrief-only: CE'. Each side shows 'none' if empty."""
    co = ", ".join(on_closeout_only) if on_closeout_only else "none"
    db = ", ".join(on_debrief_only) if on_debrief_only else "none"
    return f"closeout-only: {co}; debrief-only: {db}"


def _disc_record(location, tail, date, discrepancy, program):
    return {
        _DISC_LOCATION: location,
        _DISC_TAIL:     tail,
        _DISC_DATE:     date,
        _DISC_VALUE:    discrepancy,
        _DISC_PROGRAM:  program,
    }


def build_discrepancy_records(report):
    """Flatten a reconcile() report into a list of discrepancy records (one per
    discrepancy). Returns [] for skipped runs or runs with no findings. Double
    debriefs are included even on the clean-email path, since duplicate_debriefs
    is populated independently of has_discrepancies."""
    if report.get("skipped"):
        return []
    # Strip any '-Program' suffix so the table only ever sees the bare 3-letter
    # airport code (e.g. 'SGF-Envoy' -> 'SGF', 'CVG' stays 'CVG').
    location = report.get("location", "").strip().upper().split("-")[0]
    date = report.get("date", "")
    records = []
    for fleet, info in report.get("fleets", {}).items():
        d = info.get("discrepancies", {})
        for x in d.get("missing_in_debrief", []):
            records.append(_disc_record(location, x["tail"], date,
                                        "Missing from Debrief", fleet))
        for x in d.get("missing_in_closeout", []):
            records.append(_disc_record(location, x["tail"], date,
                                        "Missing from Closeout", fleet))
        for x in d.get("service_mismatches", []):
            disc = (f"Service Mismatch "
                    f"({_fmt_service_delta(x['on_closeout_only'], x['on_debrief_only'])})")
            records.append(_disc_record(location, x["tail"], date, disc, fleet))
        for x in d.get("probable_typos", []):
            # Report the closeout-entered (typo'd) tail; name the likely-correct one.
            disc = (f"Possible Typo (entered '{x['closeout_tail']}', "
                    f"likely '{x['debrief_tail']}')")
            if x.get("service_match") is False:
                disc += (f"; services also differ "
                         f"({_fmt_service_delta(x.get('on_closeout_only', []), x.get('on_debrief_only', []))})")
            records.append(_disc_record(location, x["closeout_tail"], date, disc, fleet))
        for x in info.get("duplicate_debriefs", []):
            records.append(_disc_record(location, x["tail"], date,
                                        "Double Debrief", fleet))
    return records


def post_discrepancy_records(records):
    """POST each discrepancy record to the Power Automate webhook as its own HTTP
    request (one row per discrepancy, mirroring PSA Tail Add). Best-effort: a
    failed POST is logged and the remaining records still go out. Returns the
    count actually accepted by the flow."""
    import requests

    if not records:
        return 0
    if not DISCREPANCY_WEBHOOK_URL:
        print("[DISCREPANCY_WEBHOOK_URL not set — discrepancy records not posted]",
              flush=True)
        return 0

    sent = 0
    for rec in records:
        try:
            resp = requests.post(
                DISCREPANCY_WEBHOOK_URL,
                headers={"Content-Type": "application/json"},
                json=rec, timeout=30)
            if resp.status_code not in (200, 201, 202):
                print(f"[discrepancy POST failed ({resp.status_code}) for "
                      f"{rec[_DISC_PROGRAM]} {rec[_DISC_TAIL]}: {resp.text[:200]}]",
                      flush=True)
                continue
            sent += 1
        except Exception as e:  # noqa: BLE001 — best-effort; keep posting the rest
            print(f"[discrepancy POST error for {rec[_DISC_PROGRAM]} "
                  f"{rec[_DISC_TAIL]}: {e}]", flush=True)
    return sent


# ----- TEXT REPORT (no-API fallback / logging) -------------------------------

def format_report(report):
    if report.get("skipped"):
        return f"SKIPPED: {report['reason']} (location={report.get('location')}, date={report.get('date')})"
    lines = [f"Reconciliation — {report['location']} {report['date']} "
             f"(submitted by {report['submitter']})"]
    if not report["has_discrepancies"] and not report.get("has_duplicates"):
        lines.append("  No discrepancies. Closeout and debriefs are in agreement.")
        return "\n".join(lines)
    for fleet, info in report["fleets"].items():
        d = info["discrepancies"]
        dups = info.get("duplicate_debriefs", [])
        if not any(d.values()) and not dups:
            continue
        lines.append(f"\n  [{fleet}]  closeout tails={info['closeout_tail_count']}, "
                     f"debrief tails={info['debrief_tail_count']}")
        for x in d["missing_in_debrief"]:
            lines.append(f"    - {x['tail']}: on CLOSEOUT, missing from DEBRIEF "
                         f"(services: {', '.join(x['closeout_services']) or 'none'})")
        for x in d["missing_in_closeout"]:
            lines.append(f"    - {x['tail']}: in DEBRIEF, missing from CLOSEOUT "
                         f"(services: {', '.join(x['debrief_services']) or 'none'})")
        for x in d.get("probable_typos", []):
            note = (f"    - {x['closeout_tail']}: on CLOSEOUT, no debrief match "
                    f"-- PROBABLE TYPO of debrief tail {x['debrief_tail']}")
            if x.get("service_match") is False:
                note += (f"; services also differ "
                         f"(closeout-only={x['on_closeout_only'] or '[]'}, "
                         f"debrief-only={x['on_debrief_only'] or '[]'})")
            lines.append(note)
        for x in d["service_mismatches"]:
            lines.append(f"    - {x['tail']}: service mismatch  "
                         f"closeout-only={x['on_closeout_only'] or '[]'}, "
                         f"debrief-only={x['on_debrief_only'] or '[]'}")
        for x in dups:
            occ = "; ".join("[" + ", ".join(o) + "]" if o else "[none]"
                            for o in x["occurrences"])
            lines.append(f"    - {x['tail']}: DOUBLE DEBRIEF — submitted on "
                         f"{x['count']} debriefs (services per submission: {occ})")
    return "\n".join(lines)


def _load_payload():
    """Get the closeout webhook payload for a GitHub Actions run.

    Resolution order:
      1. CLI arg: path to a JSON file (local testing).
      2. CLOSEOUT_PAYLOAD env var containing the JSON string (set by the workflow
         from the repository_dispatch client_payload).
      3. CLOSEOUT_PAYLOAD_FILE env var: path to a JSON file.

    Unwraps to the closeout body dict (the object whose keys are "6", "4", "27"...)
    regardless of how it is nested. GitHub's repository_dispatch caps client_payload
    at 10 top-level properties, so the flow wraps the body under a single
    "closeout" key; this also accepts a "body" wrapper or a bare body.
    """
    import sys
    raw = None
    if len(sys.argv) > 1:
        with open(sys.argv[1]) as fh:
            raw = fh.read()
    elif os.environ.get("CLOSEOUT_PAYLOAD"):
        raw = os.environ["CLOSEOUT_PAYLOAD"]
    elif os.environ.get("CLOSEOUT_PAYLOAD_FILE"):
        with open(os.environ["CLOSEOUT_PAYLOAD_FILE"]) as fh:
            raw = fh.read()
    else:
        raise SystemExit("No payload provided (CLI arg, CLOSEOUT_PAYLOAD, "
                         "or CLOSEOUT_PAYLOAD_FILE).")
    data = json.loads(raw)
    # Peel known wrapper keys until we reach the dict that holds the closeout
    # fields. Order matters: client_payload -> closeout -> body.
    for key in ("client_payload", "closeout", "body"):
        while isinstance(data, dict) and key in data and isinstance(data[key], dict):
            data = data[key]
    return data


def _construct_sender_from_submitter(submitter):
    """Build a ('First Last', 'first.last@foxtrotaviation.com') pair from the
    JotForm field 3 submitter name. Returns (None, None) if the name is empty
    or doesn't yield any usable local-part tokens.

    Splits on whitespace, lowercases, drops empties, and joins with '.'. So
    'Sam Kosco' -> ('Sam Kosco', 'sam.kosco@foxtrotaviation.com'). Multi-word
    names (middle names, suffixes) come through as 'first.middle.last@...';
    if that bounces the dispatch falls back to the configured default."""
    name = (submitter or "").strip()
    parts = [p for p in name.lower().split() if p]
    if not parts:
        return None, None
    return name, f"{'.'.join(parts)}@foxtrotaviation.com"


def _iah_dispatch_already_sent(date):
    """True if the IAH dispatch email has already been sent for this closeout
    date (field 4) on a prior run — i.e. this submission is a resubmission for a
    day we've already dispatched. Best-effort: if state can't be read we return
    False so the dispatch still goes out (better a duplicate than a silent miss)."""
    if not date:
        return False
    state = _graph_get_json(IAH_DISPATCH_STATE_PATH) or {}
    return date in (state.get("dispatched_dates") or [])


def _record_iah_dispatch(date):
    """Remember that the IAH dispatch email was sent for this closeout date so
    future resubmissions for the same day skip it. Keeps a bounded recent-date
    history. Best-effort: a write failure is logged but doesn't fail the run."""
    if not date:
        return
    state = _graph_get_json(IAH_DISPATCH_STATE_PATH) or {}
    dates = [d for d in (state.get("dispatched_dates") or []) if d != date]
    dates.append(date)  # most recent last
    state["dispatched_dates"] = dates[-IAH_DISPATCH_STATE_KEEP:]
    try:
        _graph_put_json(IAH_DISPATCH_STATE_PATH, state)
        print(f"\n[IAH dispatch date recorded: {date}]", flush=True)
    except Exception as e:  # noqa: BLE001 — don't fail the run over state bookkeeping
        print(f"\n[IAH dispatch state write failed for {date}: {e}]", flush=True)


def _send_iah_dispatch(body):
    """Build and send the IAH dispatch email. For IAH submissions, the sender
    address and sign-off name are constructed from field 3 (submitter name).
    If that send fails, the email is rebuilt and resent from the configured
    IAH_DISPATCH_SENDER_EMAIL / IAH_DISPATCH_SENDER_NAME defaults so the
    sign-off always matches the actual sender. Honors SEND_EMAIL=false.

    Returns True if the email was actually sent via Graph, False if it was only
    drafted (SEND_EMAIL=false). Raises if every sender attempt failed."""
    submitter = (body.get("3") or "").strip()
    primary_name, primary_email = _construct_sender_from_submitter(submitter)

    # Build the attempt list. Skip the primary if it would be identical to the
    # default (e.g. submitter IS David Blatt) — no point trying the same address
    # twice.
    attempts = []
    if primary_name and (primary_name, primary_email) != (
            IAH_DISPATCH_SENDER_NAME, IAH_DISPATCH_SENDER_EMAIL):
        attempts.append((primary_name, primary_email, "submitter"))
    attempts.append((IAH_DISPATCH_SENDER_NAME, IAH_DISPATCH_SENDER_EMAIL, "default"))

    send_on = os.environ.get("SEND_EMAIL", "true").lower() == "true"

    last_error = None
    for i, (name, email_from, label) in enumerate(attempts):
        dispatch = build_iah_dispatch_email(body, sender_name=name, sender_email=email_from)
        print(f"\n----- IAH DISPATCH EMAIL ({label}) -----", flush=True)
        print(f"To: {', '.join(dispatch['to'])}\nFrom: {dispatch['from']}\n"
              f"Subject: {dispatch['subject']}\n", flush=True)
        if i == 0:
            print(dispatch["body"], flush=True)  # log full body only on first attempt

        if not send_on:
            print(f"\n[SEND_EMAIL=false — {label} draft only, not sent]", flush=True)
            return False

        try:
            send_email_via_graph(dispatch)
            print(f"\n[IAH dispatch sent via Graph ({label}: {email_from})]", flush=True)
            return True
        except Exception as e:  # noqa: BLE001 — retry as default on any send failure
            last_error = e
            print(f"\n[IAH dispatch send failed for {label} ({email_from}): {e}]",
                  flush=True)
            if i < len(attempts) - 1:
                print(f"[Falling back to default sender ...]", flush=True)

    # Both attempts exhausted — re-raise so the outer try/except in main() logs it.
    raise RuntimeError(f"IAH dispatch failed for all senders; last error: {last_error}")


def main():
    body = _load_payload()

    # IAH-only: send the per-closeout dispatch email independently of
    # reconciliation. Wrapped so a failure here never blocks reconciliation.
    # The dispatch is once-per-day: if we already sent it for this closeout's
    # date, this submission is a resubmission (edited values) — skip the dispatch
    # but still run the comparison below.
    loc_base = (body.get("6") or "").strip().upper().split("-")[0]
    if loc_base == "IAH":
        iah_date = (body.get("4") or "").strip()
        if _iah_dispatch_already_sent(iah_date):
            print(f"\n[IAH dispatch skipped — already sent for {iah_date} "
                  f"(resubmission); comparison still runs]", flush=True)
        else:
            try:
                if _send_iah_dispatch(body):
                    _record_iah_dispatch(iah_date)
            except Exception as e:  # noqa: BLE001
                print(f"\n[IAH dispatch email failed: {e}]", flush=True)

    report = reconcile(body)

    # Always log a human-readable report to the Actions console.
    print(format_report(report), flush=True)

    if report.get("skipped"):
        return  # location skipped (DFW/STL AD HOC) or unparseable — no email

    send_on = os.environ.get("SEND_EMAIL", "true").lower() == "true"

    if report.get("has_discrepancies"):
        email = draft_discrepancy_email(report)   # Claude API
    else:
        email = build_clean_email(report)         # deterministic, no API
    if email is None:
        return

    # Log the draft so it's visible even if sending fails.
    print("\n----- EMAIL -----", flush=True)
    print(f"To: {', '.join(email['to'])}\nSubject: {email['subject']}\n", flush=True)
    print(email["body"], flush=True)

    if send_on:
        send_email_via_graph(email)
        print("\n[sent via Graph]", flush=True)
    else:
        print("\n[SEND_EMAIL=false — draft only, not sent]", flush=True)

    # Discrepancy records → Power Automate (one POST per discrepancy → one Excel
    # row). Built from the same report and runs on every non-skipped submission,
    # including clean-email runs whose only finding is a double debrief. Honors
    # SEND_EMAIL=false (draft-only: records are logged but not posted).
    records = build_discrepancy_records(report)
    if records:
        print(f"\n----- DISCREPANCY RECORDS ({len(records)}) -----", flush=True)
        for rec in records:
            print(json.dumps(rec), flush=True)
        if send_on:
            sent = post_discrepancy_records(records)
            print(f"\n[{sent}/{len(records)} discrepancy records posted to PA]",
                  flush=True)
        else:
            print("\n[SEND_EMAIL=false — discrepancy records drafted only, not posted]",
                  flush=True)


if __name__ == "__main__":
    main()
